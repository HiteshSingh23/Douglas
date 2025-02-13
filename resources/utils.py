import openpyxl

def read_xlsx_column_by_name(file_path, column_name, sheet_index=0):
    """
    Reads data from a specific column in an XLSX file by its header name.
    :param file_path: Path to the XLSX file.
    :param column_name: Name of the column to extract.
    :param sheet_index: Index of the sheet (default is the first sheet, 0-based).
    :return: List of values from the specified column.
    """
    values = []
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.worksheets[sheet_index]  # Get the sheet by index

    # Find the column index by header
    header_row = next(sheet.iter_rows(min_row=1, max_row=1))  # Get the first row (header)
    column_index = None
    for idx, cell in enumerate(header_row, start=1):
        if cell.value == column_name:
            column_index = idx
            break

    if column_index is None:
        raise ValueError(f"Column '{column_name}' not found in the header row.")

    # Read data from the found column (skip the header row)
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=column_index, max_col=column_index):
        for cell in row:
            values.append(cell.value)

    workbook.close()
    return values


